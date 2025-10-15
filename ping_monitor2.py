import subprocess
import platform
import socket
import re
import os
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime


def get_local_ip():
    """Get the default IP address of the computer"""
    try:
        # Create a socket to determine the default IP
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        local_ip = s.getsockname()[0]
        s.close()
        return local_ip
    except Exception:
        return "Unable to determine"


def ping_host(ip_address, count=3):
    """
    Ping a host multiple times and return detailed statistics
    Returns: dict with status, ping_times, packet_loss, min, max, avg
    """
    param = "-n" if platform.system().lower() == "windows" else "-c"
    command = ["ping", param, str(count), ip_address]

    result = {
        'status': False,
        'ping_times': [],
        'packet_loss': 100.0,
        'packets_sent': count,
        'packets_received': 0,
        'min_time': None,
        'max_time': None,
        'avg_time': None
    }

    try:
        output = subprocess.check_output(
            command,
            stderr=subprocess.STDOUT,
            universal_newlines=True,
            timeout=15
        )

        # Extract individual ping times
        if platform.system().lower() == "windows":
            # Windows format: "time=Xms" or "time<1ms"
            times = re.findall(r'time[=<](\d+)ms', output)
            result['ping_times'] = [int(t) for t in times]

            # Extract packet loss
            loss_match = re.search(r'\((\d+)%\s+loss\)', output)
            if loss_match:
                result['packet_loss'] = int(loss_match.group(1))

            # Extract min/max/avg from statistics
            stats_match = re.search(r'Minimum = (\d+)ms, Maximum = (\d+)ms, Average = (\d+)ms', output)
            if stats_match:
                result['min_time'] = int(stats_match.group(1))
                result['max_time'] = int(stats_match.group(2))
                result['avg_time'] = int(stats_match.group(3))
        else:
            # Linux/Mac format: "time=X.X ms"
            times = re.findall(r'time=([\d.]+)\s*ms', output)
            result['ping_times'] = [float(t) for t in times]

            # Extract packet loss
            loss_match = re.search(r'(\d+)%\s+packet loss', output)
            if loss_match:
                result['packet_loss'] = int(loss_match.group(1))

            # Extract min/max/avg from statistics
            stats_match = re.search(r'min/avg/max[^=]*=\s*([\d.]+)/([\d.]+)/([\d.]+)', output)
            if stats_match:
                result['min_time'] = float(stats_match.group(1))
                result['avg_time'] = float(stats_match.group(2))
                result['max_time'] = float(stats_match.group(3))

        result['packets_received'] = len(result['ping_times'])
        result['status'] = result['packets_received'] > 0

        # Calculate stats if not found in output
        if result['ping_times']:
            if result['min_time'] is None:
                result['min_time'] = min(result['ping_times'])
            if result['max_time'] is None:
                result['max_time'] = max(result['ping_times'])
            if result['avg_time'] is None:
                result['avg_time'] = sum(result['ping_times']) / len(result['ping_times'])

        return result

    except (subprocess.CalledProcessError, subprocess.TimeoutExpired):
        return result


def save_to_excel(results, local_ip):
    """Save ping results to Excel file on Desktop - overwrites if file exists"""
    # Get Desktop path
    desktop = Path.home() / "Desktop"
    excel_path = desktop / "ping_results.xlsx"

    # Try to delete existing file if it exists
    if excel_path.exists():
        try:
            excel_path.unlink()
        except PermissionError:
            # File is open, try alternative name with timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            excel_path = desktop / f"ping_results_{timestamp}.xlsx"
            print(f"Warning: Original file is open. Saving as: {excel_path.name}")

    # Create new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ping Results"

    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    cell_alignment = Alignment(horizontal="center", vertical="center")

    # Add title and info
    ws['A1'] = "Ping Test Results"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Test Date/Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A3'] = f"Local IP Address: {local_ip}"

    # Create headers
    row = 5
    headers = ["Host Name", "IP Address", "Status", "Packets Sent", "Packets Received",
               "Packet Loss (%)", "Min (ms)", "Max (ms)", "Avg (ms)", "All Ping Times"]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row, col_num, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = cell_alignment

    start_row = 6

    # Add data rows - one row per IP
    row = start_row
    for host_name, ip_address, ping_result in results:
        # Host Name
        ws.cell(row, 1, host_name)
        ws.cell(row, 1).alignment = Alignment(horizontal="left", vertical="center")

        # IP Address
        ws.cell(row, 2, ip_address)
        ws.cell(row, 2).alignment = cell_alignment

        # Status - Show TRUE/FALSE
        status_value = ping_result['status']  # Boolean True/False
        status_cell = ws.cell(row, 3, status_value)
        status_cell.alignment = cell_alignment
        # Color code: Green for True, Red for False
        if status_value:
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            status_cell.font = Font(color="006100", bold=True)
        else:
            status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            status_cell.font = Font(color="9C0006", bold=True)

        # Packets Sent
        ws.cell(row, 4, ping_result['packets_sent'])
        ws.cell(row, 4).alignment = cell_alignment

        # Packets Received
        ws.cell(row, 5, ping_result['packets_received'])
        ws.cell(row, 5).alignment = cell_alignment

        # Packet Loss
        loss_cell = ws.cell(row, 6, f"{ping_result['packet_loss']}%")
        loss_cell.alignment = cell_alignment
        # Color code packet loss
        if ping_result['packet_loss'] == 0:
            loss_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif ping_result['packet_loss'] < 100:
            loss_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        else:
            loss_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Min Time
        min_val = ping_result['min_time'] if ping_result['min_time'] is not None else "N/A"
        ws.cell(row, 7, min_val)
        ws.cell(row, 7).alignment = cell_alignment

        # Max Time
        max_val = ping_result['max_time'] if ping_result['max_time'] is not None else "N/A"
        ws.cell(row, 8, max_val)
        ws.cell(row, 8).alignment = cell_alignment

        # Avg Time
        avg_val = ping_result['avg_time'] if ping_result['avg_time'] is not None else "N/A"
        if isinstance(avg_val, (int, float)):
            avg_val = round(avg_val, 2)
        ws.cell(row, 9, avg_val)
        ws.cell(row, 9).alignment = cell_alignment

        # All Ping Times
        ping_times_str = ", ".join([str(t) for t in ping_result['ping_times']]) if ping_result['ping_times'] else "N/A"
        ws.cell(row, 10, ping_times_str)
        ws.cell(row, 10).alignment = cell_alignment

        row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 20

    # Add borders to data area
    from openpyxl.styles import Border, Side
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row_num in range(start_row, row):
        for col_num in range(1, 11):
            ws.cell(row_num, col_num).border = thin_border

    # Save workbook
    wb.save(excel_path)
    return excel_path


def main():
    # Define hosts to ping
    hosts = [
        ("EMS-Server", "192.168.1.25"),
        ("SAP-Server", "10.10.254.202"),
        ("ERP-Server", "192.168.1.3"),
        ("PTCL-Router", "192.168.1.41"),
        ("HO-Internet", "192.168.1.96"),
        ("Internet-41", "192.168.1.98"),
        ("AV-Server", "192.168.2.36")
    ]

    # Get local IP
    local_ip = get_local_ip()

    print("=" * 80)
    print("PING TEST UTILITY")
    print("=" * 80)
    print(f"Local IP Address: {local_ip}")
    print(f"Test Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 80)
    print()

    # Store results
    results = []

    # Ping each host
    for host_name, ip_address in hosts:
        print(f"Pinging {host_name} ({ip_address}) - 3 times...", end=" ")
        ping_result = ping_host(ip_address, count=3)

        results.append((host_name, ip_address, ping_result))

        # Display result
        status_text = "Reachable" if ping_result['status'] else "Unreachable"
        packet_loss_text = f"{ping_result['packet_loss']}%"

        print(f"Status = {status_text}, Packet Loss = {packet_loss_text}")

        if ping_result['ping_times']:
            times_str = ", ".join([f"{t}ms" for t in ping_result['ping_times']])
            print(f"  → Ping Times: {times_str}")
            print(
                f"  → Min: {ping_result['min_time']}ms, Max: {ping_result['max_time']}ms, Avg: {round(ping_result['avg_time'], 2)}ms")
        else:
            print(f"  → All packets lost")
        print()

    print("=" * 80)
    print("SUMMARY")
    print("=" * 80)
    print(f"{'Host':<12} {'IP Address':<18} {'Status':<12} {'Loss':<8} {'Min':<8} {'Max':<8} {'Avg':<8}")
    print("-" * 80)

    for host_name, ip_address, ping_result in results:
        status_text = "Reachable" if ping_result['status'] else "Unreachable"
        loss_text = f"{ping_result['packet_loss']}%"
        min_text = f"{ping_result['min_time']}ms" if ping_result['min_time'] is not None else "N/A"
        max_text = f"{ping_result['max_time']}ms" if ping_result['max_time'] is not None else "N/A"
        avg_text = f"{round(ping_result['avg_time'], 2)}ms" if ping_result['avg_time'] is not None else "N/A"

        print(
            f"{host_name:<12} {ip_address:<18} {status_text:<12} {loss_text:<8} {min_text:<8} {max_text:<8} {avg_text:<8}")

    print("=" * 80)
    print()

    # Save to Excel
    try:
        excel_path = save_to_excel(results, local_ip)
        print(f"Results saved to: {excel_path}")
        if "ping_results.xlsx" in str(excel_path):
            print("Note: Previous data has been replaced with current test results.")
        print()
    except Exception as e:
        print(f"Error saving to Excel: {e}")
        print("Please close the Excel file if it's open and try again.")
        print()

    # Wait for user input
    input("Press any key to exit...")


if __name__ == "__main__":
    main()