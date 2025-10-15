#!/usr/bin/env python3
"""
Enhanced Network Scanner Script with Hostname Resolution and Critical IP Testing
================================================================================
A cross-platform Python script that scans a subnet using CIDR notation,
pings each IP address, resolves hostnames for active hosts, tests connectivity
to critical infrastructure IPs, and exports results to an Excel file.

Requirements:
- Python 3.6+
- openpyxl library
- ipaddress library (built-in)
- subprocess library (built-in)
- concurrent.futures library (built-in)
- socket library (built-in)

Author: Claude
"""

import ipaddress
import subprocess
import platform
import concurrent.futures
import time
import os
import socket
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import sys

# Define critical infrastructure IPs to test connectivity
CRITICAL_IPS = [
    ("192.168.1.25", "EMS"),
    ("10.10.254.202", "SAP"),
    ("192.168.1.3", "IP-3"),
    ("192.168.1.41", "IP-41"),
    ("192.168.1.96", "IP-96"),
    ("192.168.1.98", "IP-98"),
    ("192.168.2.36", "IP-36")
]


def resolve_hostname(ip_str, timeout=2):
    """
    Resolve hostname for a given IP address using reverse DNS lookup.

    Args:
        ip_str (str): IP address as string
        timeout (int): Timeout in seconds for DNS resolution

    Returns:
        str: Hostname if resolved, empty string if not resolvable
    """
    try:
        # Set socket timeout for DNS resolution
        socket.setdefaulttimeout(timeout)

        # Perform reverse DNS lookup
        hostname = socket.gethostbyaddr(ip_str)[0]

        # Clean up hostname (remove domain suffix if present and long)
        if len(hostname) > 50:
            hostname_parts = hostname.split('.')
            if len(hostname_parts) > 1:
                hostname = hostname_parts[0]  # Take just the first part

        return hostname

    except (socket.herror, socket.gaierror, socket.timeout, OSError):
        # DNS resolution failed
        return ""
    except Exception:
        # Any other error
        return ""
    finally:
        # Reset socket timeout to default
        socket.setdefaulttimeout(None)


def get_computer_info(ip_str, timeout=5):
    """
    Get computer name and username information for a given IP address.
    Uses multiple methods depending on the operating system.

    Args:
        ip_str (str): IP address as string
        timeout (int): Timeout in seconds for each method

    Returns:
        tuple: (computer_name, username) - empty strings if not retrievable
    """
    computer_name = ""
    username = ""

    try:
        system = platform.system().lower()

        # Method 1: Try NetBIOS name resolution (Windows networks)
        computer_name = get_netbios_name(ip_str, timeout)

        # Method 2: Try to get logged-in user information
        if computer_name:  # Only try username if we have computer name
            username = get_remote_username(ip_str, computer_name, timeout)

        # Method 3: If NetBIOS failed, try other methods
        if not computer_name:
            if system == "windows":
                computer_name = get_computer_name_windows(ip_str, timeout)
            else:
                computer_name = get_computer_name_unix(ip_str, timeout)

        # Clean up names
        computer_name = computer_name.strip() if computer_name else ""
        username = username.strip() if username else ""

        # Limit length to prevent Excel display issues
        computer_name = computer_name[:50] if computer_name else ""
        username = username[:50] if username else ""

    except Exception as e:
        # Silently handle errors - we don't want to crash the scan
        pass

    return computer_name, username


def get_netbios_name(ip_str, timeout=3):
    """
    Get NetBIOS computer name using nbtstat (Windows) or nmblookup (Linux/Mac).

    Args:
        ip_str (str): IP address as string
        timeout (int): Timeout in seconds

    Returns:
        str: Computer name or empty string
    """
    try:
        system = platform.system().lower()

        if system == "windows":
            # Use nbtstat on Windows
            cmd = ["nbtstat", "-A", ip_str]
        else:
            # Try nmblookup on Unix-like systems (part of samba-common)
            cmd = ["nmblookup", "-A", ip_str]

        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=timeout
        )

        if result.returncode == 0:
            output = result.stdout.lower()

            if system == "windows":
                # Parse Windows nbtstat output
                lines = result.stdout.split('\n')
                for line in lines:
                    if '<00>' in line and 'unique' in line.lower():
                        # Extract computer name from NetBIOS table
                        parts = line.strip().split()
                        if parts:
                            name = parts[0].strip()
                            # Filter out obvious non-computer names
                            if name and not name.startswith('__') and len(name) > 1:
                                return name
            else:
                # Parse Unix nmblookup output
                lines = result.stdout.split('\n')
                for line in lines:
                    if '<00>' in line and 'unique' in line.lower():
                        # Extract computer name
                        match = re.search(r'(\S+)\s+<00>', line)
                        if match:
                            return match.group(1)

    except (subprocess.TimeoutExpired, FileNotFoundError, Exception):
        pass

    return ""


def get_remote_username(ip_str, computer_name, timeout=3):
    """
    Try to get the currently logged-in username on a remote computer.
    Uses multiple methods for better success rate.

    Args:
        ip_str (str): IP address as string
        computer_name (str): Computer name
        timeout (int): Timeout in seconds

    Returns:
        str: Username or empty string
    """
    try:
        system = platform.system().lower()

        if system == "windows":
            # Method 1: Try WMI query first (most reliable)
            try:
                cmd = ["wmic", "/node:" + ip_str, "computersystem", "get", "username", "/value"]
                result = subprocess.run(
                    cmd,
                    capture_output=True,
                    text=True,
                    timeout=timeout
                )

                if result.returncode == 0 and result.stdout:
                    for line in result.stdout.split('\n'):
                        if 'username=' in line.lower():
                            username = line.split('=', 1)[1].strip()
                            if username and username.lower() != 'username':
                                # Extract just the username part (remove domain)
                                if '\\' in username:
                                    username = username.split('\\')[-1]
                                return username

            except (subprocess.TimeoutExpired, FileNotFoundError):
                pass

            # Method 2: Try NetBIOS session enumeration
            try:
                cmd = ["nbtstat", "-A", ip_str]
                result = subprocess.run(
                    cmd,
                    capture_output=True,
                    text=True,
                    timeout=timeout
                )

                if result.returncode == 0:
                    lines = result.stdout.split('\n')
                    for line in lines:
                        # Look for <03> type entries which indicate logged-in users
                        if '<03>' in line and 'unique' in line.lower():
                            parts = line.strip().split()
                            if parts and len(parts[0]) > 0:
                                potential_user = parts[0].strip()
                                # Filter out computer names and group names
                                if potential_user != computer_name and not potential_user.endswith('__MSBROWSE__'):
                                    return potential_user

            except (subprocess.TimeoutExpired, FileNotFoundError):
                pass

            # Method 3: Try 'query user' command
            try:
                cmd = ["query", "user", "/server:" + ip_str]
                result = subprocess.run(
                    cmd,
                    capture_output=True,
                    text=True,
                    timeout=timeout
                )

                if result.returncode == 0:
                    lines = result.stdout.split('\n')
                    for line in lines[1:]:  # Skip header
                        if line.strip():
                            # Parse the query user output
                            parts = line.split()
                            if len(parts) >= 1 and parts[0] not in ['USERNAME', '>']:
                                # First column is username
                                username = parts[0].replace('>', '').strip()
                                if username:
                                    return username

            except (subprocess.TimeoutExpired, FileNotFoundError):
                pass

            # Method 4: Try PowerShell with CIM
            try:
                ps_cmd = f'(Get-CimInstance -ComputerName {ip_str} -ClassName Win32_ComputerSystem -ErrorAction SilentlyContinue).UserName'
                cmd = ["powershell", "-Command", ps_cmd]
                result = subprocess.run(
                    cmd,
                    capture_output=True,
                    text=True,
                    timeout=timeout + 1
                )

                if result.returncode == 0 and result.stdout.strip():
                    username = result.stdout.strip()
                    if '\\' in username:
                        username = username.split('\\')[-1]
                    return username

            except (subprocess.TimeoutExpired, FileNotFoundError):
                pass

        else:
            # For Unix-like systems, limited options without SSH
            pass

    except Exception:
        pass

    return ""


def get_computer_name_windows(ip_str, timeout=3):
    """
    Get computer name using Windows-specific methods.

    Args:
        ip_str (str): IP address as string
        timeout (int): Timeout in seconds

    Returns:
        str: Computer name or empty string
    """
    try:
        # Try ping with -a flag to resolve name
        cmd = ["ping", "-a", "-n", "1", "-w", str(timeout * 1000), ip_str]
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=timeout + 2
        )

        if result.returncode == 0:
            # Look for computer name in ping output
            lines = result.stdout.split('\n')
            for line in lines:
                if f'pinging {ip_str}' in line.lower() or f'ping {ip_str}' in line.lower():
                    continue
                if 'pinging' in line.lower() and '[' + ip_str + ']' in line:
                    # Extract name from "Pinging computername [IP]"
                    match = re.search(r'pinging\s+(\S+)\s+\[', line, re.IGNORECASE)
                    if match:
                        name = match.group(1)
                        if name != ip_str:  # Make sure it's not just the IP
                            return name

    except (subprocess.TimeoutExpired, Exception):
        pass

    return ""


def get_computer_name_unix(ip_str, timeout=3):
    """
    Get computer name using Unix-specific methods.

    Args:
        ip_str (str): IP address as string
        timeout (int): Timeout in seconds

    Returns:
        str: Computer name or empty string
    """
    try:
        # Try using host command for reverse DNS
        cmd = ["host", ip_str]
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=timeout
        )

        if result.returncode == 0:
            # Parse host command output
            match = re.search(r'domain name pointer\s+(\S+)', result.stdout)
            if match:
                hostname = match.group(1).rstrip('.')
                # Extract just the computer name part
                return hostname.split('.')[0]

    except (subprocess.TimeoutExpired, FileNotFoundError, Exception):
        pass

    return ""


def test_connectivity_to_critical_ips(source_ip, timeout=2):
    """
    Test connectivity from a source IP to all critical infrastructure IPs.

    Args:
        source_ip (str): Source IP address
        timeout (int): Timeout for each ping test

    Returns:
        dict: Dictionary with target IPs as keys and status as values
    """
    results = {}

    system = platform.system().lower()

    for target_ip, label in CRITICAL_IPS:
        try:
            if system == "windows":
                cmd = ["ping", "-n", "1", "-w", str(timeout * 1000), target_ip]
            else:
                cmd = ["ping", "-c", "1", "-W", str(timeout), target_ip]

            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                timeout=timeout + 2
            )

            stdout_lower = result.stdout.lower()

            # Check if ping was successful
            if system == "windows":
                success = (f"reply from {target_ip}:" in stdout_lower and
                           "bytes=" in stdout_lower and "time=" in stdout_lower)
            else:
                success = (f"bytes from {target_ip}" in stdout_lower and
                           "time=" in stdout_lower)

            if success or result.returncode == 0:
                results[f"{target_ip} ({label})"] = "✓ Reachable"
            else:
                results[f"{target_ip} ({label})"] = "✗ Unreachable"

        except (subprocess.TimeoutExpired, Exception):
            results[f"{target_ip} ({label})"] = "✗ Timeout"

    return results


def ping_host(ip_str, debug=False):
    """
    Ping a single host and return its status with precise response parsing.

    Args:
        ip_str (str): IP address as string
        debug (bool): Enable debug output for troubleshooting

    Returns:
        tuple: (ip_address, status) where status is 'Active', 'Host Unreachable', or 'Request Timeout'
    """
    try:
        # Determine ping command based on operating system
        system = platform.system().lower()

        if system == "windows":
            # Windows ping command - single ping with 3 second timeout
            cmd = ["ping", "-n", "1", "-w", "3000", ip_str]
        else:
            # Linux/Mac ping command - single ping with 3 second timeout
            cmd = ["ping", "-c", "1", "-W", "3", ip_str]

        # Execute ping command
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=6
        )

        # Get original output (preserve case for better matching)
        stdout_original = result.stdout.strip()
        stderr_original = result.stderr.strip()
        stdout_lower = stdout_original.lower()
        stderr_lower = stderr_original.lower()

        if debug:
            print(f"\n=== DEBUG - {ip_str} ===")
            print(f"Return code: {result.returncode}")
            print(f"STDOUT: '{stdout_original}'")
            print(f"STDERR: '{stderr_original}'")

        # === STEP 1: Check for ACTIVE responses (successful ping) ===
        # These patterns definitively indicate an active, responding host

        if system == "windows":
            # Windows success patterns - must contain the target IP
            active_conditions = [
                # Standard successful reply
                f"reply from {ip_str}:" in stdout_lower and "bytes=" in stdout_lower and "time=" in stdout_lower,
                # Alternative format
                f"reply from {ip_str}" in stdout_lower and "ttl=" in stdout_lower
            ]
        else:
            # Linux/Mac success patterns - must contain the target IP
            active_conditions = [
                # Standard successful reply
                f"64 bytes from {ip_str}:" in stdout_lower and "time=" in stdout_lower,
                # Alternative format
                f"bytes from {ip_str}" in stdout_lower and "icmp_seq=" in stdout_lower,
                # Another common format
                f"ping: 56 data bytes" in stdout_lower and f"64 bytes from {ip_str}" in stdout_lower
            ]

        # Check if any active condition is met
        if any(active_conditions):
            if debug:
                print("Status: ACTIVE - Successful ping response detected")
            return (ip_str, "Active")

        # === STEP 2: Check for HOST UNREACHABLE responses ===
        # These patterns indicate the IP is not assigned or routing issues

        unreachable_conditions = []

        if system == "windows":
            unreachable_conditions = [
                "destination host unreachable" in stdout_lower,
                "destination net unreachable" in stdout_lower,
                "general failure" in stdout_lower,
                "transmit failed" in stdout_lower,
                "hardware error" in stdout_lower,
                "no resources" in stdout_lower
            ]
        else:
            unreachable_conditions = [
                "destination host unreachable" in stdout_lower,
                "network unreachable" in stdout_lower,
                "no route to host" in stdout_lower,
                "connect: network is unreachable" in stdout_lower,
                "connect: no route to host" in stdout_lower,
                "network is down" in stdout_lower,
                "host is down" in stdout_lower
            ]

        # Check if any unreachable condition is met
        if any(unreachable_conditions):
            if debug:
                print("Status: HOST UNREACHABLE - Network/routing issue detected")
            return (ip_str, "Host Unreachable")

        # === STEP 3: Check for TIMEOUT responses ===
        # These patterns indicate the IP might be assigned but host is not responding

        timeout_conditions = []

        if system == "windows":
            timeout_conditions = [
                "request timed out" in stdout_lower,
                "request timeout" in stdout_lower,
                # No reply but no explicit unreachable message
                result.returncode != 0 and "unreachable" not in stdout_lower and "failure" not in stdout_lower
            ]
        else:
            timeout_conditions = [
                "100% packet loss" in stdout_lower,
                "0 received" in stdout_lower and "transmitted" in stdout_lower,
                "no answer" in stdout_lower,
                "no reply" in stdout_lower,
                # Pattern for when packet is sent but no response
                "1 packets transmitted, 0 received" in stdout_lower,
                "1 packets transmitted, 0 packets received" in stdout_lower
            ]

        # Check if any timeout condition is met
        if any(timeout_conditions):
            if debug:
                print("Status: REQUEST TIMEOUT - Host not responding to ping")
            return (ip_str, "Request Timeout")

        # === STEP 4: Fallback analysis based on return code and content ===

        if result.returncode == 0:
            # Return code 0 should mean success, but we didn't find success patterns
            # This might be a parsing issue - be conservative and call it active
            if debug:
                print("Status: ACTIVE (fallback) - Return code 0 but unusual output")
            return (ip_str, "Active")

        elif result.returncode == 1:
            # Return code 1 - analyze content more carefully
            if "unreachable" in stdout_lower or "failure" in stdout_lower:
                if debug:
                    print("Status: HOST UNREACHABLE (fallback) - Return code 1 with unreachable indicators")
                return (ip_str, "Host Unreachable")
            else:
                # Likely timeout scenario
                if debug:
                    print("Status: REQUEST TIMEOUT (fallback) - Return code 1, likely timeout")
                return (ip_str, "Request Timeout")

        elif result.returncode == 2:
            # Return code 2 usually means network/host unreachable
            if debug:
                print("Status: HOST UNREACHABLE (fallback) - Return code 2")
            return (ip_str, "Host Unreachable")

        else:
            # Any other return code - default to timeout
            if debug:
                print(f"Status: REQUEST TIMEOUT (fallback) - Unknown return code {result.returncode}")
            return (ip_str, "Request Timeout")

    except subprocess.TimeoutExpired:
        if debug:
            print("Status: REQUEST TIMEOUT - Process timeout")
        return (ip_str, "Request Timeout")
    except FileNotFoundError:
        print(f"Error: ping command not found for {ip_str}")
        return (ip_str, "Request Timeout")
    except Exception as e:
        if debug:
            print(f"Status: REQUEST TIMEOUT - Exception: {e}")
        return (ip_str, "Request Timeout")


def ping_and_gather_info(ip_str, debug=False, test_critical=True):
    """
    Ping a host and gather all available information if active.

    Args:
        ip_str (str): IP address as string
        debug (bool): Enable debug output for troubleshooting
        test_critical (bool): Test connectivity to critical IPs

    Returns:
        tuple: (ip_address, status, hostname, computer_name, username, critical_connectivity)
    """
    # First ping the host
    ip, status = ping_host(ip_str, debug)

    # Initialize info variables
    hostname = ""
    computer_name = ""
    username = ""
    critical_connectivity = {}

    # Only gather additional info for active hosts
    if status == "Active":
        # Get hostname via DNS
        hostname = resolve_hostname(ip_str)

        # Get computer name and username
        computer_name, username = get_computer_info(ip_str)

        # Test connectivity to critical infrastructure IPs
        if test_critical:
            critical_connectivity = test_connectivity_to_critical_ips(ip_str)

        if debug:
            if hostname:
                print(f"Resolved hostname for {ip_str}: {hostname}")
            if computer_name:
                print(f"Found computer name for {ip_str}: {computer_name}")
            if username:
                print(f"Found username for {ip_str}: {username}")
            if critical_connectivity:
                print(f"Critical connectivity results for {ip_str}:")
                for target, status in critical_connectivity.items():
                    print(f"  {target}: {status}")

    return (ip, status, hostname, computer_name, username, critical_connectivity)


def validate_cidr(cidr_input):
    """
    Validate CIDR notation input.

    Args:
        cidr_input (str): User input for CIDR notation

    Returns:
        ipaddress.IPv4Network or None: Valid network object or None if invalid
    """
    try:
        network = ipaddress.IPv4Network(cidr_input, strict=False)
        return network
    except (ipaddress.AddressValueError, ipaddress.NetmaskValueError, ValueError):
        return None


def scan_network(network, enable_debug=False):
    """
    Scan all IP addresses in the given network with full information gathering.

    Args:
        network (ipaddress.IPv4Network): Network to scan
        enable_debug (bool): Enable debug output for first few IPs

    Returns:
        list: List of tuples containing (ip_address, status, hostname, computer_name, username, critical_connectivity)
    """
    print(f"Scanning network: {network}")
    print(f"Total hosts to scan: {network.num_addresses}")

    # Get list of all IP addresses in the network
    ip_list = [str(ip) for ip in network.hosts()]

    # If it's a single host network (like /32), include the network address
    if network.num_addresses == 1:
        ip_list = [str(network.network_address)]

    results = []

    # Use ThreadPoolExecutor for concurrent scanning
    print("Starting comprehensive network scan (ping + hostname + computer info + critical connectivity)...")

    # For debugging, test first few IPs with debug enabled
    if enable_debug and len(ip_list) > 0:
        print("\n=== DEBUG MODE: Testing first 3 IPs ===")
        for i, ip in enumerate(ip_list[:3]):
            result = ping_and_gather_info(ip, debug=True, test_critical=True)
            results.append(result)
            print(f"Result: {result}")
        print("=== END DEBUG ===\n")

        # Continue with remaining IPs without debug
        remaining_ips = ip_list[3:]
    else:
        remaining_ips = ip_list

    with concurrent.futures.ThreadPoolExecutor(max_workers=15) as executor:
        # Reduced max_workers to account for additional connectivity tests
        # Submit all remaining scan tasks
        future_to_ip = {executor.submit(ping_and_gather_info, ip, False, True): ip for ip in remaining_ips}

        # Process completed tasks
        completed = len(results)  # Account for debug results
        for future in concurrent.futures.as_completed(future_to_ip):
            ip_str, status, hostname, computer_name, username, critical_connectivity = future.result()
            results.append((ip_str, status, hostname, computer_name, username, critical_connectivity))
            completed += 1

            # Show progress with detailed info for active hosts
            if status == "Active":
                info_parts = []
                if hostname:
                    info_parts.append(f"DNS: {hostname}")
                if computer_name:
                    info_parts.append(f"Computer: {computer_name}")
                if username:
                    info_parts.append(f"User: {username}")

                # Show critical connectivity summary
                if critical_connectivity:
                    reachable = sum(1 for v in critical_connectivity.values() if "Reachable" in v)
                    info_parts.append(f"Critical IPs: {reachable}/{len(CRITICAL_IPS)} reachable")

                if info_parts:
                    info_str = " | ".join(info_parts)
                    print(f"Progress: {completed}/{len(ip_list)} - Found: {ip_str} ({info_str})")
                else:
                    print(f"Progress: {completed}/{len(ip_list)} - Found: {ip_str} (Active - no additional info)")
            elif completed % 15 == 0 or completed == len(ip_list):
                print(f"Progress: {completed}/{len(ip_list)} hosts scanned")

    # Sort results by IP address
    results.sort(key=lambda x: ipaddress.IPv4Address(x[0]))
    return results


def get_desktop_path():
    """
    Get the path to the user's desktop directory across different operating systems.

    Returns:
        str: Path to desktop directory
    """
    system = platform.system().lower()

    if system == "windows":
        # Windows desktop path
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        # Alternative path if the above doesn't exist
        if not os.path.exists(desktop):
            desktop = os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop")
        if not os.path.exists(desktop):
            # Fallback to user home directory
            desktop = os.path.expanduser("~")
    elif system == "darwin":  # macOS
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        if not os.path.exists(desktop):
            desktop = os.path.expanduser("~")
    else:  # Linux and other Unix-like systems
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        # Some Linux distributions might have localized desktop names
        if not os.path.exists(desktop):
            # Try common alternatives
            alternatives = [
                os.path.join(os.path.expanduser("~"), "Escritorio"),  # Spanish
                os.path.join(os.path.expanduser("~"), "Bureau"),  # French
                os.path.join(os.path.expanduser("~"), "Schreibtisch"),  # German
            ]
            for alt in alternatives:
                if os.path.exists(alt):
                    desktop = alt
                    break
            else:
                desktop = os.path.expanduser("~")  # Fallback to home

    return desktop


def export_to_excel(results, filename="Network Scanner with Critical IP Testing.xlsx"):
    """
    Export comprehensive scan results to an Excel file on the desktop.

    Args:
        results (list): List of tuples containing (ip_address, status, hostname, computer_name, username, critical_connectivity)
        filename (str): Output filename for Excel file
    """
    # Get desktop path and create full file path
    desktop_path = get_desktop_path()
    full_path = os.path.join(desktop_path, filename)

    print(f"Exporting results to {full_path}...")

    # Create new workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Network Scan Results"

    # Set up headers - include critical IP columns
    headers = ["IP Address", "Status", "Hostname (DNS)", "Computer Name", "Username"]

    # Add columns for each critical IP
    for target_ip, label in CRITICAL_IPS:
        headers.append(f"{label}\n({target_ip})")

    ws.append(headers)

    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Add data rows and count statuses
    active_count = 0
    unreachable_count = 0
    timeout_count = 0
    resolved_hostnames = 0
    found_computer_names = 0
    found_usernames = 0

    for ip, status, hostname, computer_name, username, critical_connectivity in results:
        # Display values or "N/A" if empty
        hostname_display = hostname if hostname else "N/A"
        computer_display = computer_name if computer_name else "N/A"
        username_display = username if username else "N/A"

        # Build row data
        row_data = [ip, status, hostname_display, computer_display, username_display]

        # Add critical connectivity results
        if critical_connectivity:
            for target_ip, label in CRITICAL_IPS:
                key = f"{target_ip} ({label})"
                connectivity_status = critical_connectivity.get(key, "N/A")
                # Simplify display - just show Reachable/Unreachable/Timeout/N/A
                if "Reachable" in connectivity_status:
                    row_data.append("✓ Yes")
                elif "Unreachable" in connectivity_status:
                    row_data.append("✗ No")
                elif "Timeout" in connectivity_status:
                    row_data.append("⏱ Timeout")
                else:
                    row_data.append("N/A")
        else:
            # No connectivity data (host not active)
            for _ in CRITICAL_IPS:
                row_data.append("N/A")

        ws.append(row_data)

        # Count statuses and information
        if status == "Active":
            active_count += 1
            if hostname:
                resolved_hostnames += 1
            if computer_name:
                found_computer_names += 1
            if username:
                found_usernames += 1
        elif status == "Host Unreachable":
            unreachable_count += 1
        elif status == "Request Timeout":
            timeout_count += 1

        # Color code status column and information columns
        row_num = ws.max_row
        status_cell = ws[f"B{row_num}"]
        hostname_cell = ws[f"C{row_num}"]
        computer_cell = ws[f"D{row_num}"]
        username_cell = ws[f"E{row_num}"]

        if status == "Active":
            # Green for active hosts
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            status_cell.font = Font(color="006100")

            # Highlight information cells if data is available
            if hostname:
                hostname_cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                hostname_cell.font = Font(color="006100")
            if computer_name:
                computer_cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                computer_cell.font = Font(color="006100")
            if username:
                username_cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                username_cell.font = Font(color="006100")

            # Color code critical connectivity cells
            for i, (target_ip, label) in enumerate(CRITICAL_IPS):
                col_letter = chr(70 + i)  # F, G, H, I, J, K, L (starting from column F)
                conn_cell = ws[f"{col_letter}{row_num}"]

                if "✓ Yes" in str(conn_cell.value):
                    # Green for reachable
                    conn_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    conn_cell.font = Font(color="006100")
                elif "✗ No" in str(conn_cell.value):
                    # Red for unreachable
                    conn_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    conn_cell.font = Font(color="9C0006")
                elif "⏱ Timeout" in str(conn_cell.value):
                    # Orange for timeout
                    conn_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    conn_cell.font = Font(color="9C6500")

        elif status == "Host Unreachable":
            # Red for unreachable hosts
            status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            status_cell.font = Font(color="9C0006")
        elif status == "Request Timeout":
            # Orange/yellow for timeout hosts
            status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            status_cell.font = Font(color="9C6500")

    # Add summary at the bottom
    ws.append([])  # Empty row
    summary_start_row = ws.max_row + 1
    ws.append(["Summary:", "", "", "", ""])
    ws.append([f"Total Scanned:", len(results), "", "", ""])
    ws.append([f"Active Hosts:", active_count, "", "", ""])
    ws.append([f"Hostnames Resolved:", resolved_hostnames, "", "", ""])
    ws.append([f"Computer Names Found:", found_computer_names, "", "", ""])
    ws.append([f"Usernames Found:", found_usernames, "", "", ""])
    ws.append([f"Host Unreachable:", unreachable_count, "", "", ""])
    ws.append([f"Request Timeout:", timeout_count, "", "", ""])

    # Add scan timestamp
    ws.append([f"Scan Date:", time.strftime("%Y-%m-%d %H:%M:%S"), "", "", ""])

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save the workbook
    try:
        wb.save(full_path)
        print(f"Results successfully exported to Desktop: {filename}")
        return True
    except PermissionError:
        print(f"Error: Cannot write to {full_path}. File may be open in another application.")
        # Try saving to current directory as fallback
        try:
            wb.save(filename)
            print(f"Fallback: Results saved to current directory: {filename}")
            return True
        except:
            return False
    except Exception as e:
        print(f"Error saving file: {e}")
        # Try saving to current directory as fallback
        try:
            wb.save(filename)
            print(f"Fallback: Results saved to current directory: {filename}")
            return True
        except:
            return False


def main():
    """
    Main function to run the comprehensive network scanner.
    """
    print("=" * 80)
    print("         COMPREHENSIVE NETWORK SCANNER")
    print("    (IP Address | Hostname | Computer Name | Username | Critical IPs)")
    print("=" * 80)
    print()

    # Get CIDR input from user
    while True:
        cidr_input = input("Enter subnet in CIDR notation (e.g., 192.168.1.0/24): ").strip()

        if not cidr_input:
            print("Please enter a valid CIDR notation.")
            continue

        network = validate_cidr(cidr_input)
        if network is None:
            print("Invalid CIDR notation. Please try again.")
            print("Examples: 192.168.1.0/24, 10.0.0.0/16, 172.16.0.0/12")
            continue

        # Check if network is too large
        if network.num_addresses > 512:
            response = input(
                f"Warning: This will scan {network.num_addresses} addresses and gather detailed info. This may take a while. Continue? (y/n): ")
            if response.lower() not in ['y', 'yes']:
                continue

        break

    print()
    print("Note: This scan will gather comprehensive information including:")
    print("  • Ping status")
    print("  • DNS hostname resolution")
    print("  • NetBIOS computer names")
    print("  • Currently logged-in usernames (where possible)")
    print("  • Connectivity tests to critical infrastructure IPs:")
    for target_ip, label in CRITICAL_IPS:
        print(f"      - {target_ip} ({label})")
    print()
    print("This process may take longer than a basic ping scan.")
    print("Some information may require administrative privileges on target systems.")
    print()

    # Record start time
    start_time = time.time()

    try:
        # Ask user if they want debug mode for troubleshooting
        if network.num_addresses <= 10:
            debug_response = input("Enable debug mode to see ping details? (y/n): ")
            enable_debug = debug_response.lower() in ['y', 'yes']
        else:
            enable_debug = False

        # Scan the network
        results = scan_network(network, enable_debug)

        # Calculate scan time
        scan_time = time.time() - start_time

        # Count results
        active_hosts = [r for r in results if r[1] == 'Active']
        resolved_hosts = [r for r in active_hosts if r[2]]  # Has hostname
        computer_hosts = [r for r in active_hosts if r[3]]  # Has computer name
        user_hosts = [r for r in active_hosts if r[4]]  # Has username

        print(f"\nScan completed in {scan_time:.2f} seconds")
        print(f"Found {len(active_hosts)} active hosts")
        print(f"Resolved hostnames for {len(resolved_hosts)} hosts")
        print(f"Found computer names for {len(computer_hosts)} hosts")
        print(f"Found usernames for {len(user_hosts)} hosts")
        print(f"Found {sum(1 for r in results if r[1] == 'Host Unreachable')} unreachable hosts")
        print(f"Found {sum(1 for r in results if r[1] == 'Request Timeout')} timeout hosts")

        # Show active hosts with detailed info
        if active_hosts:
            print(f"\nActive hosts with detailed information:")
            for ip, status, hostname, computer_name, username, critical_connectivity in active_hosts:
                info_parts = []
                if hostname:
                    info_parts.append(f"DNS: {hostname}")
                if computer_name:
                    info_parts.append(f"Computer: {computer_name}")
                if username:
                    info_parts.append(f"User: {username}")

                # Show critical connectivity summary
                if critical_connectivity:
                    reachable = sum(1 for v in critical_connectivity.values() if "Reachable" in v)
                    info_parts.append(f"Critical Connectivity: {reachable}/{len(CRITICAL_IPS)}")

                if info_parts:
                    info_str = " | ".join(info_parts)
                    print(f"  {ip} - {info_str}")
                else:
                    print(f"  {ip} - Active (no additional info available)")

                # Show detailed connectivity breakdown
                if critical_connectivity:
                    for target, conn_status in critical_connectivity.items():
                        status_symbol = "✓" if "Reachable" in conn_status else "✗"
                        print(f"      {status_symbol} {target}: {conn_status}")

        # Export to Excel
        if export_to_excel(results):
            print("\nScan completed successfully!")
        else:
            print("\nScan completed but there was an error exporting to Excel.")

    except KeyboardInterrupt:
        print("\n\nScan interrupted by user.")
        sys.exit(1)
    except Exception as e:
        print(f"\nAn error occurred: {e}")
        sys.exit(1)


if __name__ == "__main__":
    # Check if openpyxl is installed
    try:
        import openpyxl
    except ImportError:
        print("Error: openpyxl library is not installed.")
        print("Please install it using: pip install openpyxl")
        sys.exit(1)

    main()